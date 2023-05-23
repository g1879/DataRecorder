# -*- coding:utf-8 -*-
from pathlib import Path

from openpyxl.reader.excel import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.workbook import Workbook

from .tools import process_content, ok_list


class OriginalSetter(object):
    def __init__(self, recorder):
        self._recorder = recorder

    def cache_size(self, size):
        """设置缓存大小
        :param size: 缓存大小
        :return: None
        """
        if not isinstance(size, int) or size < 0:
            raise TypeError('cache_size值只能是int，且必须>=0')
        self._recorder._cache = size

    def path(self, path, file_type=None):
        """设置文件路径
        :param path: 文件路径
        :param file_type: 要设置的文件类型，为空则从文件名中获取
        :return: None
        """
        if file_type is not None and isinstance(file_type, str):
            self._recorder._type = file_type
        elif isinstance(path, str):
            self._recorder._type = path.split('.')[-1].lower()
        elif isinstance(path, Path):
            self._recorder._type = path.suffix[1:].lower()
        else:
            raise TypeError(f'参数path只能是str或Path，非{type(path)}。')

        if self._recorder._type not in self._recorder.SUPPORTS and 'any' not in self._recorder.SUPPORTS:
            raise TypeError(f'只支持{"、".join(self._recorder.SUPPORTS)}格式文件。')

        if self._recorder._path:
            self._recorder.record()  # 更换文件前自动记录剩余数据

        self._recorder._path = str(path) if isinstance(path, Path) else path

    def file_type(self, file_type):
        """指定文件类型，无视文件后缀名"""
        self._recorder._type = file_type


class BaseSetter(OriginalSetter):
    def table(self, name):
        """设置默认表名
        :param name: 表名
        :return: None
        """
        self._recorder._table = name

    def before(self, before):
        """设置在数据前面补充的列
        :param before: 列表、元组或字符串，为字符串时则补充一列
        :return: None
        """
        self._recorder.record()

        if before is None:
            self._recorder._before = []
        elif isinstance(before, (list, dict)):
            self._recorder._before = before
        elif isinstance(before, tuple):
            self._recorder._before = list(before)
        else:
            self._recorder._before = [before]

    def after(self, after):
        """设置在数据后面补充的列
        :param after: 列表、元组或字符串，为字符串时则补充一列
        :return: None
        """
        self._recorder.record()

        if after is None:
            self._recorder._after = []
        elif isinstance(after, (list, dict)):
            self._recorder._after = after
        elif isinstance(after, tuple):
            self._recorder._after = list(after)
        else:
            self._recorder._after = [after]

    def head(self, head):
        """设置表头。只有 csv 和 xlsx 格式支持设置表头
        :param head: 表头，列表或元组
        :return: None
        """
        if self._recorder.type == 'xlsx':
            set_xlsx_head(self._recorder.path, head, self._recorder.table)

        elif self._recorder.type == 'csv':
            set_csv_head(self._recorder.path, head, self._recorder.encoding, self._recorder.delimiter,
                         self._recorder.quote_char)

        else:
            raise TypeError('只能对xlsx和csv文件设置表头。')

    def encoding(self, encoding):
        """设置编码
        :param encoding: 编码格式
        :return: None
        """
        self._recorder._encoding = encoding


class FillerSetter(BaseSetter):
    def sign(self, value):
        """设置sign值
        :param value: 筛选条件文本
        :return: None
        """
        self._recorder._sign = value

    def deny_sign(self, on_off=True):
        """设置是否反向匹配sign
        :param on_off: bool表示开或关
        :return: None
        """
        self._recorder._deny_sign = on_off

    def key_cols(self, cols):
        """设置作为关键字的列，可以是多列
        :param cols: 列号或列名，或它们组成的list或tuple
        :return: None
        """
        if cols is True:
            self._recorder._key_cols = True
        elif isinstance(cols, int) and cols > 0:
            self._recorder._key_cols = [cols]
        elif isinstance(cols, str):
            self._recorder._key_cols = [int(cols)] if cols.isdigit() else [column_index_from_string(cols)]
        elif isinstance(cols, (list, tuple)):
            self._recorder._key_cols = [i if isinstance(i, int) and i > 0 else
                                        int(i) if i.isdigit() else column_index_from_string(i) for i in cols]
        else:
            raise TypeError('col值只能是int或str，且必须大于0。')

    def sign_col(self, col):
        """设置用于判断是否已填数据的列
        :param col: 列号或列名
        :return: None
        """
        if col is True or (isinstance(col, int) and col > 0):
            self._recorder._sign_col = col
        elif isinstance(col, str):
            self._recorder._sign_col = int(col) if col.isdigit() else column_index_from_string(col)
        else:
            raise TypeError('col值只能是True、int或str，且必须大于0。')

    def data_col(self, col):
        """设置用于填充数据的列
        :param col: 列号或列名
        :return: None
        """
        if isinstance(col, int) and col > 0:
            self._recorder._data_col = col
        elif isinstance(col, str):
            self._recorder._data_col = column_index_from_string(col)
        else:
            raise TypeError('col值只能是int或str，且必须大于0。')

    def begin_row(self, row):
        """设置数据开始的行
        :param row: 行号
        :return: None
        """
        if not isinstance(row, int) or row < 1:
            raise TypeError('row值只能是int，且必须大于0')
        self._recorder._begin_row = row

    def path(self, path, key_cols=None, begin_row=None, sign_col=None,
             data_col=None, sign=None, deny_sign=None):
        """设置文件路径
        :param path: 保存的文件路径
        :param key_cols: 作为关键字的列，可以是多列
        :param begin_row: 数据开始的行，默认表头一行
        :param sign_col: 用于判断是否已填数据的列
        :param data_col: 要填入数据的第一列
        :param sign: 按这个值判断是否已填数据
        :param deny_sign: 是否反向匹配sign，即筛选指不是sign的行
        """
        if not path or not Path(path).exists():
            raise FileNotFoundError('文件不存在')
        super().path(path)
        self._recorder.set.key_cols(key_cols or self._recorder.key_cols)
        self._recorder.set.begin_row(begin_row or self._recorder.begin_row)
        self._recorder.set.sign_col(sign_col or self._recorder.sign_col)
        self._recorder.set.sign(sign or self._recorder.sign)
        self._recorder.set.data_col(data_col or self._recorder.data_col)
        self._recorder.set.deny_sign(deny_sign if deny_sign is not None else self._recorder.deny_sign)

    def link_style(self, style):
        """设置excel的链接样式
        :param style: Font对象
        :return: None
        """
        self._recorder._link_font = style

    def delimiter(self, delimiter):
        """设置csv文件分隔符
        :param delimiter: 分隔符
        :return: None
        """
        self._recorder._delimiter = delimiter

    def quote_char(self, quote_char):
        """设置csv文件引用符
        :param quote_char: 引用符
        :return: None
        """
        self._recorder._quote_char = quote_char


class RecorderSetter(BaseSetter):
    def follow_styles(self, on_off=True):
        """设置是否跟随最后一行的style，只有xlsx格式有效
        :param on_off: True或False
        :return: None
        """
        self._recorder._follow_styles = on_off
        if not on_off:
            self._recorder._row_styles = None
            self._recorder._row_styles_len = None

    def col_height(self, height):
        """设置行高，只有xlsx格式有效
        :param height: 行高
        :return: None
        """
        self._recorder._col_height = height

    def style(self, style):
        """设置样式，只有xlsx格式有效
        :param style: CellStyle对象
        :return: None
        """
        self._recorder._style = style

    def path(self, path, file_type=None):
        """设置文件路径
        :param path: 文件路径
        :param file_type: 要设置的文件类型，为空则从文件名中获取
        :return: None
        """
        super().path(path=path, file_type=file_type)
        self._recorder._row_styles = None

    def delimiter(self, delimiter):
        """设置csv文件分隔符
        :param delimiter: 分隔符
        :return: None
        """
        self._recorder._delimiter = delimiter

    def quote_char(self, quote_char):
        """设置csv文件引用符
        :param quote_char: 引用符
        :return: None
        """
        self._recorder._quote_char = quote_char


class DBSetter(BaseSetter):
    def path(self, path, file_type=None):
        """重写父类方法
        :param path: 文件路径
        :param file_type: 文件类型
        :return: None
        """
        super().path(path, file_type)
        if self._recorder._conn is not None:
            self._recorder._close_connection()
        self._recorder._connect()


def set_csv_head(file_path, head, encoding='utf-8', delimiter=',', quote_char='"'):
    """设置csv文件的表头
    :param file_path: 文件路径
    :param head: 表头列表或元组
    :param encoding: 编码
    :param delimiter: 分隔符
    :param quote_char: 引用符
    :return: None
    """
    from csv import writer
    if Path(file_path).exists():
        with open(file_path, 'r', newline='', encoding=encoding) as f:
            content = "".join(f.readlines()[1:])

        with open(file_path, 'w', newline='', encoding=encoding) as f:
            csv_write = writer(f, delimiter=delimiter, quotechar=quote_char)
            csv_write.writerow(ok_list(head))

        with open(file_path, 'a+', newline='', encoding=encoding) as f:
            f.write(f'{content}')

    else:
        with open(file_path, 'w', newline='', encoding=encoding) as f:
            csv_write = writer(f, delimiter=delimiter, quotechar=quote_char)
            csv_write.writerow(ok_list(head))


def set_xlsx_head(file_path, head, table):
    """设置xlsx文件的表头
    :param file_path: 文件路径
    :param head: 表头列表或元组
    :param table: 工作表名称
    :return: None
    """
    wb = load_workbook(file_path) if Path(file_path).exists() else Workbook()
    if table:
        ws = wb[table] if table in [i.title for i in wb.worksheets] else wb.create_sheet(title=table)
    else:
        ws = wb.active

    for key, i in enumerate(head, 1):
        ws.cell(1, key).value = process_content(i, True)

    wb.save(file_path)
    wb.close()
