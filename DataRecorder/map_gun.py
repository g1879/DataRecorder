# -*- coding:utf-8 -*-
from csv import reader as csv_reader, writer as csv_writer
from pathlib import Path
from typing import Union

from openpyxl import load_workbook, Workbook

from .base import BaseRecorder, _parse_coord, _process_content


class MapGun(BaseRecorder):
    """把二维数据填充到以左上角坐标为起点的范围"""

    def __init__(self, path: Union[str, Path] = None,
                 coord: Union[str, tuple, list] = None,
                 float_coord: bool = True):
        """初始化                                              \n
        :param path: 保存的文件路径
        :param coord: 目标左上角坐标
        :param float_coord: 保存数据后坐标是否移动到底部
        """
        super().__init__(path, 1)
        self.coord = coord or [1, 1]
        self.float_coord = float_coord

    @property
    def coord(self) -> list:
        """起始坐标"""
        return self._coord

    @coord.setter
    def coord(self, coord: Union[str, tuple, list]) -> None:
        """设置填写坐标                                                               \n
        :param coord: 接受几种形式：'A3', '3,1', (3, 1), [3, 1]，除第一种外都是行在前
        :return: None
        """
        self._coord = list(_parse_coord(coord, None, int))

    @property
    def cache_size(self) -> int:
        """返回缓存大小"""
        return self._cache

    @cache_size.setter
    def cache_size(self, cache_size: int) -> None:
        """固定缓存大小                   \n
        :param cache_size: 缓存大小
        :return: None
        """
        pass

    def add_data(self, data: Union[list, tuple], coord: Union[str, tuple, list] = None) -> None:
        """接收二维数据，若是一维的，每个元素作为一行看待    \n
        :param data: 二维数据
        :param coord: 左上角坐标
        :return: None
        """
        if coord is not None:
            self.coord = coord
        self._data = data
        self.record()

    def _record(self) -> None:
        """记录数据"""
        if self.type == 'xlsx':
            self._to_xlsx()
        elif self.type == 'csv':
            self._to_csv()

        if self.float_coord:
            self.coord[0] += len(self._data)

    def _to_xlsx(self) -> None:
        """记录数据到xlsx文件"""
        wb = load_workbook(self.path) if Path(self.path).exists() else Workbook()
        ws = wb.active

        row, col = self.coord
        for i in self._data:
            if not isinstance(i, (list, tuple)):
                i = (i,)

            for ind, j in enumerate(self._data_to_list(i)):
                ws.cell(row, col + ind).value = _process_content(j, True)

            row += 1

        wb.save(self.path)
        wb.close()

    def _to_csv(self) -> None:
        """填写数据到xlsx文件"""
        if not Path(self.path).exists():
            with open(self.path, 'w', encoding=self.encoding):
                pass

        with open(self.path, 'r', encoding=self.encoding) as f:
            reader = csv_reader(f, delimiter=self.delimiter, quotechar=self.quote_char)
            lines = list(reader)
            lines_len = len(lines)
            row, col = self.coord

            # 若行数不够，填充行数
            for _ in range(row + len(self._data) - lines_len):
                lines.append([])

            # 填入数据
            for ind, i in enumerate(self._data):
                if not isinstance(i, (list, tuple)):
                    i = (i,)

                now_data = self._data_to_list(i)
                row_num = row + ind - 1

                # 若列数不够，填充空列
                lines[row_num].extend([None] * (col - len(lines[row_num]) + len(now_data) - 1))

                # 填充一行数据
                for k, j in enumerate(now_data):
                    lines[row_num][col + k - 1] = _process_content(j)

            writer = csv_writer(open(self.path, 'w', encoding=self.encoding, newline=''),
                                delimiter=self.delimiter, quotechar=self.quote_char)
            writer.writerows(lines)
