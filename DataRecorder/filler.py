# -*- coding:utf-8 -*-
from csv import reader as csv_reader, writer as csv_writer
from pathlib import Path
from time import sleep

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

from .base import BaseRecorder
from .setter import FillerSetter
from .style.cell_style import CellStyle, NoneStyle
from .tools import parse_coord, get_usable_coord, process_content, data_to_list_or_dict


class Filler(BaseRecorder):
    def __init__(self, path=None, cache_size=None, key_cols=True, begin_row=2,
                 sign_col=True, data_col=None, sign=None, deny_sign=False):
        """用于处理表格文件的工具
        :param path: 保存的文件路径
        :param cache_size: 每接收多少条记录写入文件，传入0表示不自动保存
        :param key_cols: 作为关键字的列，可以是多列，可以是列编号或序号，从1开始，True表示获取整行
        :param begin_row: 数据开始的行，默认表头一行
        :param sign_col: 用于判断是否已填数据的列，可以是列编号或序号，从1开始，True表示获取所有行，不进行判断
        :param data_col: 要填入数据的第一列，从1开始
        :param sign: 按这个值筛选需要的行纳入keys
        :param deny_sign: 是否反向匹配sign，即筛选sign_col列值不是sign的行
        """
        super().__init__(None, cache_size)
        self._delimiter = ','  # csv文件分隔符
        self._quote_char = '"'  # csv文件引用符
        self._key_cols = None
        self._begin_row = None
        self._sign_col = None
        self._data_col = None
        self._sign = None
        self._deny_sign = False
        self.row_num_title = 'row'
        if not data_col:
            data_col = sign_col if sign_col else 1
        self.set.path(path, key_cols, begin_row, sign_col, data_col, sign, deny_sign)
        s = CellStyle()
        s.font.set_color("0000FF")
        s.font.set_underline('single')
        self._link_style = s

    @property
    def sign(self):
        """按这个值筛选需要的行纳入keys"""
        return self._sign

    @property
    def deny_sign(self):
        """返回是否反向匹配sign"""
        return self._deny_sign

    @property
    def key_cols(self):
        """返回作为关键字的列或列的集合"""
        return self._key_cols

    @property
    def sign_col(self):
        """返回用于判断是否已填数据的列"""
        return self._sign_col

    @property
    def data_col(self):
        """返回用于填充数据的列"""
        return self._data_col

    @property
    def begin_row(self):
        """返回数据开始的行号，用于获取keys，从1开始"""
        return self._begin_row

    @property
    def keys(self):
        """返回一个列表，由未执行的行数据组成。每行的格式为第一位为行号，其余为 key 列的值。
        eg.[3, '张三', 20]
        """
        if not self.path or not Path(self.path).exists():
            raise FileNotFoundError('未指定文件或文件不存在。')

        if self.type == 'csv':
            return get_csv_keys(self, False)
        elif self.type == 'xlsx':
            return get_xlsx_keys(self, False)

    @property
    def dict_keys(self):
        """返回一个列表，由未执行的行数据组成。每行的格式为dict，'row' 值为行号，其余值为第一行数据。
        如第一行数据为空，则用列号为值。如果begin_row为1，用列名作为值。
        eg.{'row': 2, 'name': '张三', 'C': '男'}
        """
        if not self.path or not Path(self.path).exists():
            raise FileNotFoundError('未指定文件或文件不存在。')

        if self.type == 'csv':
            return get_csv_keys(self, True)
        elif self.type == 'xlsx':
            return get_xlsx_keys(self, True)

    @property
    def set(self):
        """返回用于设置属性的对象"""
        if self._setter is None:
            self._setter = FillerSetter(self)
        return self._setter

    @property
    def delimiter(self):
        """返回csv文件分隔符"""
        return self._delimiter

    @property
    def quote_char(self):
        """返回csv文件引用符"""
        return self._quote_char

    def add_data(self, data, coord='newline', table=None):
        """添加数据，每次添加一行数据，可指定坐标、列号或行号
        coord只输入数字（行号）时，列号为self.data_col值，如 3；
        输入列号，或没有行号的坐标时，表示新增一行，列号为此时指定的，如'c'、',3'、(None, 3)、'None,3'；
        输入 'newline' 时，表示新增一行，列号为self.data_col值；
        输入行列坐标时，填写到该坐标，如'a3'、'3,1'、(3,1)、[3,1]；
        输入的行号列号可以是负数，代表从下往上数，-1是倒数第一行，如'a-3'、(-3, -3)
        :param data: 要添加的内容，任意格式
        :param coord: 要添加数据的坐标，可输入行号、列号或行列坐标，如'a3'、7、(3, 1)、[3, 1]、'c'
        :param table: 要写入的数据表，仅支持xlsx格式。为None表示用set.table()方法设置的值，为bool表示活动的表格
        :return: None
        """
        while self._pause_add:  # 等待其它线程写入结束
            sleep(.1)

        if not isinstance(data, (list, tuple)):
            data = (data,)

        if coord not in ('set_link', 'cover_style', 'replace_style', 'set_img', 'set_width', 'set_height'):
            coord = parse_coord(coord, self.data_col)
            if not data:
                data = ([],)
                self._data_count += 1
            # 一维数组
            elif isinstance(data, dict) or (
                    isinstance(data, (list, tuple)) and not isinstance(data[0], (list, tuple, dict))):
                data = (data_to_list_or_dict(self, data),)
                self._data_count += 1
            else:  # 二维数组
                data = [data_to_list_or_dict(self, d) for d in data]
                self._data_count += len(data)

        if self._type != 'xlsx':
            self._data.append((coord, data))

        else:
            if table is None:
                table = self._table
            elif isinstance(table, bool):
                table = None

            self._data.setdefault(table, []).append((coord, data))

        if 0 < self.cache_size <= self._data_count:
            self.record()

    def set_link(self, coord, link, content=None):
        """为单元格设置超链接
        :param coord: 单元格坐标
        :param link: 超链接，为None时删除链接
        :param content: 单元格内容
        :return: None
        """
        self.add_data((coord, link, content), 'set_link')

    def set_style(self, coord, style, replace=True):
        """为单元格设置样式
        :param coord: 单元格坐标，输入数字可设置整行，输入列名字符串可设置整列，输入'A1:C5'格式可设置指定范围
        :param style: CellStyle对象，为None则清除单元格样式
        :param replace: 是否直接替换已有样式，运行效率较高，但不能单独修改某个属性
        :return: None
        """
        s = 'replace_style' if replace else 'cover_style'
        self.add_data((coord, style), s)

    def set_img(self, coord, img_path, width=None, height=None):
        """
        :param coord: 单元格坐标，输入数字可设置整行，输入列名字符串可设置整列，输入'A1:C5'格式可设置指定范围
        :param img_path: 图片路径，为None则删除现有图片
        :param width: 图片宽
        :param height: 图片高
        :return: None
        """
        if isinstance(img_path, Path):
            img_path = str(img_path)
        self.add_data((coord, img_path, width, height), 'set_img')

    def set_row_height(self, row, height):
        """设置行高
        :param row: 行号
        :param height: 行高
        :return: None
        """
        self.add_data((row, height), 'set_height')

    def set_col_width(self, col, width):
        """设置列宽
        :param col: 列号，数字或字母
        :param width: 列宽
        :return: None
        """
        if isinstance(col, int):
            col = get_column_letter(col)
        self.add_data((col, width), 'set_width')

    def _record(self):
        """记录数据"""
        if self.type == 'xlsx':
            self._to_xlsx()
        elif self.type == 'csv':
            self._to_csv()

    def _to_xlsx(self):
        """填写数据到xlsx文件"""
        if Path(self.path).exists():
            wb = load_workbook(self.path)
            new = False
        else:
            wb = Workbook()
            new = True

        tables = [i.title for i in wb.worksheets]
        for table, table_data in self._data.items():
            if table is None:
                ws = wb.active

            else:
                if table in tables:
                    ws = wb[table]
                elif new is True:
                    new = False
                    ws = wb.active
                    ws.title = table
                    tables.append(table)
                else:
                    ws = wb.create_sheet(title=table)
                    tables.append(table)

            max_col = ws.max_column
            empty = ws.max_column == ws.max_row == 1 and ws[1][0].value is None

            for data in table_data:
                if empty:  # 如果是空文件，最大行数设为0，避免直接添加时出现空行
                    max_row = 0
                    empty = False
                else:
                    max_row = ws.max_row

                if data[0] == 'set_link':
                    coord = parse_coord(data[1][0], self.data_col)
                    row, col = get_usable_coord(coord, max_row, max_col)
                    cell = ws.cell(row, col)
                    has_link = True if cell.hyperlink else Filler
                    cell.hyperlink = None if data[1][1] is None else process_content(data[1][1], True)
                    if data[1][2] is not None:
                        cell.value = process_content(data[1][2], True)
                    if data[1][1]:
                        if self._link_style:
                            self._link_style.to_cell(cell, replace=False)
                    elif has_link:
                        NoneStyle().to_cell(cell, replace=False)
                    continue

                elif data[0] in ('replace_style', 'cover_style'):
                    mode = data[0] == 'replace_style'
                    coord = data[1][0]
                    style = NoneStyle() if data[1][1] is None else data[1][1]
                    if isinstance(coord, int) or (isinstance(coord, str) and coord.isdigit()):
                        for c in ws[coord]:
                            style.to_cell(c, replace=mode)
                        continue

                    elif isinstance(coord, str) and ':' in coord:
                        for c in ws[coord]:
                            for cc in c:
                                style.to_cell(cc, replace=mode)
                        continue

                    coord = parse_coord(coord, self.data_col)
                    row, col = get_usable_coord(coord, max_row, max_col)
                    style.to_cell(ws.cell(row, col), replace=mode)
                    continue

                elif data[0] == 'set_img':
                    coord, img_path, width, height = data[1]
                    coord = parse_coord(coord, self.data_col)
                    row, col = get_usable_coord(coord, max_row, max_col)

                    ws._images.clear()
                    if img_path is None:
                        continue

                    from openpyxl.drawing.image import Image
                    from openpyxl.utils import get_column_letter
                    img = Image(img_path)
                    if width and height:
                        img.width = width
                        img.height = height
                    elif width:
                        img.height = int(img.height * (width / img.width))
                        img.width = width
                    elif height:
                        img.width = int(img.width * (height / img.height))
                        img.height = height
                    col = get_column_letter(col)
                    ws.add_image(img, f'{col}{row}')
                    continue

                elif data[0] == 'set_width':
                    col, width = data[1]
                    ws.column_dimensions[col].width = width
                    continue

                elif data[0] == 'set_height':
                    row, height = data[1]
                    ws.row_dimensions[row].height = height
                    continue

                row, col = get_usable_coord(data[0], max_row, max_col)
                now_data = (data[1],) if not isinstance(data[1][0], (list, tuple, dict)) else data[1]

                for r, i in enumerate(now_data, row):
                    if isinstance(i, dict):
                        i = i.values()
                    for key, j in enumerate(i):
                        ws.cell(r, col + key).value = process_content(j, True)

        wb.save(self.path)
        wb.close()

    def _to_csv(self):
        """填写数据到xlsx文件"""
        if not Path(self.path).exists():
            with open(self.path, 'w', encoding=self.encoding):
                pass

        with open(self.path, 'r', encoding=self.encoding) as f:
            reader = csv_reader(f, delimiter=self.delimiter, quotechar=self.quote_char)
            lines = list(reader)
            lines_count = len(lines)

            for i in self._data:
                if i[0] == 'set_link':
                    coord = parse_coord(i[1][0], self.data_col)
                    now_data = (f'=HYPERLINK("{i[1][1]}","{i[1][2] or i[1][1]}")',)

                elif i[0] in ('cover_style', 'replace_style', 'set_img', 'set_width', 'set_height'):
                    continue

                else:
                    coord = i[0]
                    now_data = i[1]

                row, col = get_usable_coord(coord, lines_count, len(lines[0]) if lines_count else 1)
                now_data = (now_data,) if not isinstance(now_data[0], (list, tuple, dict)) else now_data

                for r, data in enumerate(now_data, row):
                    if isinstance(data, dict):
                        data = list(data.values())

                    for _ in range(r - lines_count):  # 若行数不够，填充行数
                        lines.append([])
                        lines_count += 1

                    row_num = r - 1

                    # 若列数不够，填充空列
                    lines[row_num].extend([None] * (col - len(lines[row_num]) + len(data) - 1))

                    for k, j in enumerate(data):  # 填充数据
                        lines[row_num][col + k - 1] = process_content(j)

            writer = csv_writer(open(self.path, 'w', encoding=self.encoding, newline=''), delimiter=self.delimiter,
                                quotechar=self.quote_char)
            writer.writerows(lines)


def get_xlsx_keys(filler, as_dict):
    """返回key列内容，第一位为行号，其余为key列的值
    如果as_dict为True，返回dict格式，value为第一行值，值为空或begin_row为1时用列号，'row' 值为行号
    eg.[3, '名称', 'id']
    :param filler: 记录器对象
    :param as_dict: 是否以dict格式返回数据
    :return: 关键字组成的列表或字典
    """
    wb = load_workbook(filler.path, data_only=True, read_only=True)
    if filler.table and filler.table not in [i.title for i in wb.worksheets]:
        raise RuntimeError(f'xlsx文件未包含此工作表：{filler.table}')
    ws = wb[filler.table] if filler.table else wb.active

    if ws.max_column is None:  # 遇到过read_only时无法获取列数的文件
        wb.close()
        wb = load_workbook(filler.path, data_only=True)
        ws = wb[filler.table] if filler.table else wb.active

    rows = ws.rows
    if as_dict:
        title = [filler.row_num_title]
        if filler.begin_row == 1:
            t = [get_column_letter(x) for x in range(1, ws.max_column + 1)
                 if filler.key_cols is True or x in filler.key_cols]
        else:
            try:
                t = next(rows)
                t = [x.value if x.value else get_column_letter(k) for k, x in enumerate(t, 1)
                     if filler.key_cols is True or k in filler.key_cols]

            except StopIteration:
                return []

        if len(t) != len(set(t)):
            raise RuntimeError('表头出现内容重复。')
        if filler.row_num_title in t:
            raise RuntimeError(f'表头"{filler.row_num_title}"列与行号列重复，请用set.row_num_title()方法设置不同的表头。')
        title.extend(t)

    try:
        for _ in range(filler.begin_row - 1):
            next(rows)
    except StopIteration:
        return []

    # ---------------------------------------------------------

    if filler.sign_col is True or filler.sign_col > ws.max_column:  # 获取所有行
        if filler.key_cols is True:  # 获取整行
            res = [[ind] + [i.value for i in row]
                   for ind, row in enumerate(rows, filler.begin_row)]
        else:  # 只获取对应的列
            res = [[ind] + [row[i - 1].value for i in filler.key_cols]
                   for ind, row in enumerate(rows, filler.begin_row)]

    else:  # 获取符合条件的行
        if filler.key_cols is True:  # 获取整行
            if filler.deny_sign:
                res = [[ind] + [i.value for i in row]
                       for ind, row in enumerate(rows, filler.begin_row)
                       if row[filler.sign_col - 1].value != filler.sign]
            else:
                res = [[ind] + [i.value for i in row]
                       for ind, row in enumerate(rows, filler.begin_row)
                       if row[filler.sign_col - 1].value == filler.sign]

        else:  # 只获取对应的列
            if filler.deny_sign:
                res = [[ind] + [row[i - 1].value for i in filler.key_cols]
                       for ind, row in enumerate(rows, filler.begin_row)
                       if row[filler.sign_col - 1].value != filler.sign]
            else:
                res = [[ind] + [row[i - 1].value for i in filler.key_cols]
                       for ind, row in enumerate(rows, filler.begin_row)
                       if row[filler.sign_col - 1].value == filler.sign]

    wb.close()

    if as_dict:
        res = [dict(zip(title, v)) for v in res]
    return res


def get_csv_keys(filler, as_dict):
    """返回key列内容，第一位为行号，其余为key列的值，
    如果as_dict为True，返回dict格式，value为第一行值，值为空或begin_row为1时用列号，'row'值为行号
    eg.[3, '名称', 'id']
    :param filler: 记录器对象
    :param as_dict: 是否以dict格式返回数据
    :return: 关键字组成的列表或字典
    """
    begin_row = filler.begin_row
    sign_col = filler.sign_col
    sign = '' if filler.sign is None else str(filler.sign)
    begin_row -= 1
    res = []

    with open(filler.path, 'r', encoding=filler.encoding) as f:
        reader = csv_reader(f, delimiter=filler.delimiter, quotechar=filler.quote_char)
        lines = list(reader)
        if not lines:
            return res

        if as_dict:
            title = [filler.row_num_title]
            if filler.begin_row == 1:
                t = [get_column_letter(x) for x in range(1, len(lines[0]) + 1)
                     if filler.key_cols is True or x in filler.key_cols]
            else:
                t = [x if x else get_column_letter(k) for k, x in enumerate(lines[0], 1)
                     if filler.key_cols is True or k in filler.key_cols]

            if len(t) != len(set(t)):
                raise RuntimeError('表头内容重复。')
            if filler.row_num_title in t:
                raise RuntimeError(f'表头"{filler.row_num_title}"列与行号列重复，请用set.row_num_title()方法设置不同的表头。')
            title.extend(t)

        if sign_col is not True:  # 获取符合条件的行
            sign_col -= 1
            for ind, line in enumerate(lines[begin_row:], begin_row + 1):
                row_sign = '' if sign_col > len(line) - 1 else line[sign_col]
                if row_sign != sign if filler.deny_sign else row_sign == sign:
                    if filler.key_cols is True:  # 获取整行
                        res.append([ind] + line)
                    else:  # 只获取对应的列
                        res.append([ind] + [line[i - 1] for i in filler.key_cols])

        else:  # 获取所有行
            for ind, line in enumerate(lines[begin_row:], begin_row + 1):
                if filler.key_cols is True:  # 获取整行
                    res.append([ind] + line)
                else:  # 只获取对应的列
                    res.append([ind] + [line[i - 1] for i in filler.key_cols])

    if as_dict:
        res = [dict(zip(title, v)) for v in res]
    return res
