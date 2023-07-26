# -*- coding:utf-8 -*-
from abc import abstractmethod
from pathlib import Path
from re import sub, match
from threading import Lock
from time import sleep
from typing import Union

from openpyxl import load_workbook, Workbook
from openpyxl.cell import Cell, ReadOnlyCell
from openpyxl.utils import column_index_from_string

from .tools import get_usable_path


class OriginalRecorder(object):
    """记录器的基类"""
    SUPPORTS = ('any',)

    def __init__(self, path=None, cache_size=None):
        """
        :param path: 保存的文件路径
        :param cache_size: 每接收多少条记录写入文件，0为不自动写入
        """
        self._data = []
        self._path = None
        self._type = None
        self._lock = Lock()
        self._pause_add = False  # 文件写入时暂停接收输入
        self._pause_write = False  # 标记文件正在被一个线程写入
        self.show_msg = True

        if path:
            self.set_path(path)
        self._cache = cache_size if cache_size is not None else 1000

    def __del__(self):
        """对象关闭时把剩下的数据写入文件"""
        self.record()

    @property
    def cache_size(self):
        """返回缓存大小"""
        return self._cache

    @cache_size.setter
    def cache_size(self, size):
        """设置缓存大小"""
        self.set_cache_size(size)

    def set_cache_size(self, cache_size):
        """设置缓存大小
        :param cache_size: 缓存大小
        :return: None
        """
        if not isinstance(cache_size, int) or cache_size < 0:
            raise TypeError('cache_size值只能是int，且必须>=0')
        self._cache = cache_size

    @property
    def path(self):
        """返回文件路径"""
        return self._path

    @path.setter
    def path(self, file_path):
        """设置文件路径"""
        self.set_path(file_path)

    def set_path(self, path, file_type=None):
        """设置文件路径
        :param path: 文件路径
        :param file_type: 要设置的文件类型，为空则从文件名中获取
        :return: None
        """
        if file_type is not None and isinstance(file_type, str):
            self._type = file_type
        elif isinstance(path, str):
            self._type = path.split('.')[-1].lower()
        elif isinstance(path, Path):
            self._type = path.suffix[1:].lower()
        else:
            raise TypeError(f'参数path只能是str或Path，非{type(path)}。')

        if self._type not in self.SUPPORTS and 'any' not in self.SUPPORTS:
            raise TypeError(f'只支持{"、".join(self.SUPPORTS)}格式文件。')

        if self._path:
            self.record()  # 更换文件前自动记录剩余数据

        self._path = str(path) if isinstance(path, Path) else path

    @property
    def type(self):
        """返回文件类型"""
        return self._type

    @type.setter
    def type(self, file_type):
        """设置文件类型"""
        self.set_type(file_type)

    def set_type(self, file_type):
        """指定文件类型，无视文件后缀名"""
        self._type = file_type

    @property
    def data(self):
        """返回当前保存在缓存的数据"""
        return self._data

    def record(self, new_path=None):
        """记录数据，可保存到新文件
        :param new_path: 文件另存为的路径，会保存新文件
        :return: 成功返回文件路径，失败返回未保存的数据
        """
        # 具体功能由_record()实现，本方法实现自动重试及另存文件功能
        original_path = return_path = self._path
        return_data = None
        if new_path:
            new_path = str(get_usable_path(new_path))
            return_path = self._path = new_path

            if Path(original_path).exists():
                from shutil import copy
                copy(original_path, self._path)

        if not self._data:
            return return_path

        if not self._path:
            raise ValueError('保存路径为空。')

        with self._lock:
            self._pause_add = True  # 写入文件前暂缓接收数据
            if self.show_msg:
                print(f'{self.path} 开始写入文件，切勿关闭进程')

            Path(self.path).parent.mkdir(parents=True, exist_ok=True)
            while True:
                try:
                    while self._pause_write:  # 等待其它线程写入结束
                        sleep(.1)

                    self._pause_write = True
                    self._record()
                    break

                except PermissionError:
                    if self.show_msg:
                        print('\r文件被打开，保存失败，请关闭，程序会自动重试...', end='')

                except Exception as e:
                    if self._data:
                        if self.show_msg:
                            print(f'\n{self._data}\n\n注意！！以上数据未保存')
                        return_data = self._data.copy()
                    if 'Python is likely shutting down' not in str(e):
                        raise
                    break

                finally:
                    self._pause_write = False

                sleep(.3)

            if new_path:
                self._path = original_path

            if self.show_msg:
                print(f'{self.path} 写入文件结束')
            self._data = []
            self._pause_add = False

        return return_data if return_data else return_path

    def clear(self):
        """清空缓存中的数据"""
        self._data = []

    @abstractmethod
    def add_data(self, data):
        pass

    @abstractmethod
    def _record(self):
        pass


class BaseRecorder(OriginalRecorder):
    """Recorder和Filler的父类"""
    SUPPORTS = ('xlsx', 'csv')

    def __init__(self, path=None, cache_size=None):
        """
        :param path: 保存的文件路径
        :param cache_size: 每接收多少条记录写入文件，0为不自动写入
        """
        super().__init__(path, cache_size)
        self._before = []
        self._after = []

        self._encoding = 'utf-8'
        self._delimiter = ','  # csv文件分隔符
        self._quote_char = '"'  # csv文件引用符
        self._table = None

    @property
    def before(self):
        """返回当前before内容"""
        return self._before

    @property
    def after(self):
        """返回当前after内容"""
        return self._after

    @property
    def table(self):
        """返回默认表名"""
        return self._table

    def set_table(self, table):
        """设置默认表名
        :param table: 表名
        :return: None
        """
        self._table = table

    def set_before(self, before):
        """设置在数据前面补充的列
        :param before: 列表、元组或字符串，为字符串时则补充一列
        :return: None
        """
        self.record()

        if before is None:
            self._before = []
        elif isinstance(before, (list, dict)):
            self._before = before
        elif isinstance(before, tuple):
            self._before = list(before)
        else:
            self._before = [before]

    def set_after(self, after):
        """设置在数据后面补充的列
        :param after: 列表、元组或字符串，为字符串时则补充一列
        :return: None
        """
        self.record()

        if after is None:
            self._after = []
        elif isinstance(after, (list, dict)):
            self._after = after
        elif isinstance(after, tuple):
            self._after = list(after)
        else:
            self._after = [after]

    def set_head(self, head):
        """设置表头。只有 csv 和 xlsx 格式支持设置表头
        :param head: 表头，列表或元组
        :return: None
        """
        if self.type == 'xlsx':
            _set_xlsx_head(self.path, head, self.table)

        elif self.type == 'csv':
            _set_csv_head(self.path, head, self.encoding, self.delimiter, self.quote_char)

        else:
            raise TypeError('只能对xlsx和csv文件设置表头。')

    @property
    def encoding(self):
        """返回编码格式"""
        return self._encoding

    @property
    def delimiter(self):
        """返回csv文件分隔符"""
        return self._delimiter

    @property
    def quote_char(self):
        """返回csv文件引用符"""
        return self._quote_char

    @encoding.setter
    def encoding(self, encoding):
        """设置编码格式"""
        self.set_encoding(encoding)

    @delimiter.setter
    def delimiter(self, delimiter):
        """设置csv文件分隔符"""
        self.set_delimiter(delimiter)

    @quote_char.setter
    def quote_char(self, quote_char):
        """设置csv文件引用符"""
        self.set_quote_char(quote_char)

    def set_encoding(self, encoding):
        """设置编码
        :param encoding: 编码格式
        :return: None
        """
        self._encoding = encoding

    def set_delimiter(self, delimiter):
        """设置csv文件分隔符
        :param delimiter: 分隔符
        :return: None
        """
        self._delimiter = delimiter

    def set_quote_char(self, quote_char):
        """设置csv文件引用符
        :param quote_char: 引用符
        :return: None
        """
        self._quote_char = quote_char

    @abstractmethod
    def add_data(self, data):
        pass

    @abstractmethod
    def _record(self):
        pass

    def _data_to_list(self, data):
        """将传入的数据转换为列表形式，添加前后列数据
        :param data: 要处理的数据
        :return: 转变成列表方式的数据
        """
        return_list = []
        if data is not None and not isinstance(data, (list, tuple, dict)):
            data = [data]

        for i in (self.before, data, self.after):
            if isinstance(i, dict):
                return_list.extend(list(i.values()))
            elif i is None:
                pass
            elif isinstance(i, list):
                return_list.extend(i)
            elif isinstance(i, tuple):
                return_list.extend(list(i))
            else:
                return_list.extend([str(i)])

        return return_list


def _set_csv_head(file_path: str,
                  head: Union[list, tuple],
                  encoding: str = 'utf-8',
                  delimiter: str = ',',
                  quote_char: str = '"') -> None:
    """设置csv文件的表头
    :param file_path: 文件路径
    :param head: 表头列表或元组
    :param encoding: 编码
    :param delimiter: 分隔符
    :param quote_char: 引用符
    :return: None
    """
    from csv import writer
    if Path(file_path).exists():
        with open(file_path, 'r', newline='', encoding=encoding) as f:
            content = "".join(f.readlines()[1:])

        with open(file_path, 'w', newline='', encoding=encoding) as f:
            csv_write = writer(f, delimiter=delimiter, quotechar=quote_char)
            csv_write.writerow(ok_list(head))

        with open(file_path, 'a+', newline='', encoding=encoding) as f:
            f.write(f'{content}')

    else:
        with open(file_path, 'w', newline='', encoding=encoding) as f:
            csv_write = writer(f, delimiter=delimiter, quotechar=quote_char)
            csv_write.writerow(ok_list(head))


def _set_xlsx_head(file_path: str, head: Union[list, tuple], table: str) -> None:
    """设置xlsx文件的表头
    :param file_path: 文件路径
    :param head: 表头列表或元组
    :param table: 工作表名称
    :return: None
    """
    wb = load_workbook(file_path) if Path(file_path).exists() else Workbook()
    if table:
        ws = wb[table] if table in [i.title for i in wb.worksheets] else wb.create_sheet(title=table)
    else:
        ws = wb.active

    for key, i in enumerate(head, 1):
        ws.cell(1, key).value = process_content(i, True)

    wb.save(file_path)
    wb.close()


def parse_coord(coord=None, data_col=None):
    """添加数据，每次添加一行数据，可指定坐标、列号或行号
    coord只输入数字（行号）时，列号为self.data_col值，如 3；
    输入列号，或没有行号的坐标时，表示新增一行，列号为此时指定的，如'c'、',3'、(None, 3)、'None,3'；
    输入 'newline' 时，表示新增一行，列号为self.data_col值；
    输入行列坐标时，填写到该坐标，如'a3'、'3,1'、(3,1)、[3,1]；
    输入的行号可以是负数（列号不可以），代表从下往上数，-1是倒数第一行，如'a-3'、(-3, 3)
    :param coord: 坐标、列号、行号
    :param data_col: 列号，用于只传入行号的情况
    :return: 坐标tuple：(行, 列)，或(None, 列)
    """
    return_coord = None
    if coord == 'newline':  # 新增一行，列为data_col
        return_coord = None, data_col

    elif isinstance(coord, (int, float)) and coord != 0:
        return_coord = int(coord), data_col

    elif isinstance(coord, str):
        coord = coord.replace(' ', '')

        if coord.isalpha():  # 只输入列号，要新建一行
            return_coord = None, column_index_from_string(coord)

        elif ',' in coord:  # '3,1'形式
            x, y = coord.split(',')
            if x.lower() in ('', 'new', 'none', 'newline'):
                x = None
            elif x.isdigit():
                x = int(x)
            else:
                raise ValueError('行格式不正确。')

            if y.isdigit():
                y = int(y)
            elif y.isalpha():
                y = column_index_from_string(y)
            else:
                raise TypeError('列格式不正确。')

            return_coord = x, y

        else:  # 'A3'形式
            m = match(r'^[$]?([A-Za-z]{1,3})[$]?(-?\d+)$', coord)
            if not m:
                raise ValueError('坐标格式不正确。')
            y, x = m.groups()
            return_coord = int(x), column_index_from_string(y)

    elif isinstance(coord, (tuple, list)):
        if len(coord) != 2:
            raise ValueError('coord为list或tuple时长度必须为2。')

        x = None
        if coord[0] not in (None, 'new', 'newline'):
            x = int(coord[0])

        if isinstance(coord[1], int):
            y = coord[1]
        elif isinstance(coord[1], str):
            y = column_index_from_string(coord[1])
        else:
            raise TypeError('列格式不正确。')

        return_coord = x, y

    if not return_coord or return_coord[0] == 0 or return_coord[1] == 0:
        raise ValueError(f'坐标{return_coord}格式不正确。')
    return return_coord


def process_content(content, excel=False):
    """处理单个单元格要写入的数据
    :param content: 未处理的数据内容
    :param excel: 是否为excel文件
    :return: 处理后的数据
    """
    if isinstance(content, (int, str, float, type(None))):
        data = content
    elif isinstance(content, (Cell, ReadOnlyCell)):
        data = content.value
    else:
        data = str(content)

    if excel and isinstance(data, str):
        data = sub(r'[\000-\010]|[\013-\014]|[\016-\037]', '', data)

    return data


def ok_list(data_list, excel=False, as_str=False):
    """处理列表中数据使其符合保存规范
    :param data_list: 数据列表
    :param excel: 是否保存在excel
    :param as_str: 内容是否转为字符串
    :return: 处理后的列表
    """
    if isinstance(data_list, dict):
        data_list = data_list.values()
    if as_str:
        data_list = [str(i) for i in data_list]
    return [process_content(i, excel) for i in data_list]


def get_usable_coord(coord, max_row, max_col):
    """返回真正写入文件的坐标
    :param coord: 已初步格式化的坐标，如(1, 2)、(None, 3)、(-3, -2)
    :param max_row: 文件最大行
    :param max_col: 文件最大列
    :return: 真正写入文件的坐标
    """
    row, col = coord
    if col < 0:
        col = max_col + col + 1
        if col < 1:
            raise ValueError(f'列号不能小于1。当前：{col}')

    if row is None:
        row = max_row + 1
    elif row < 0:
        row = max_row + row + 1
        if row < 1:
            raise ValueError(f'行号不能小于1。当前：{row}')

    return row, col
