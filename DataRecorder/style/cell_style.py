# -*- coding:utf-8 -*-
from copy import copy
from threading import Lock

from openpyxl.styles import Alignment, Font, Side, Border, Protection, GradientFill, PatternFill, Color


class CellStyle(object):
    font_args = ('name', 'size', 'charset', 'underline', 'color', 'scheme', 'vertAlign',
                 'bold', 'italic', 'strike', 'outline', 'shadow', 'condense', 'extend')
    border_args = ('start', 'end', 'left', 'right', 'top', 'bottom', 'diagonal', 'vertical', 'horizontal',
                   'horizontal', 'outline', 'diagonalUp', 'diagonalDown')
    alignment_args = ('horizontal', 'vertical', 'indent', 'relativeIndent', 'justifyLastLine', 'readingOrder',
                      'textRotation', 'wrapText', 'shrinkToFit')
    protection_args = ('locked', 'hidden')
    gradient_fill_args = ('type', 'degree', 'left', 'right', 'top', 'bottom', 'stop')
    pattern_fill_args = ('patternType', 'fgColor', 'bgColor')

    def __init__(self):
        """用于管理单元格样式的类"""
        self._font = None
        self._border = None
        self._alignment = None
        self._pattern_fill = None
        self._gradient_fill = None
        self._number_format = None
        self._protection = None

        # 用于覆盖目标单元格的对象
        self._Font = None
        self._Border = None
        self._Alignment = None
        self._Fill = None
        self._Protection = None

    @property
    def font(self):
        """返回用于设置单元格字体的对象"""
        if self._font is None:
            self._font = CellFont()
        return self._font

    @property
    def border(self):
        """返回用于设置单元格边框的对象"""
        if self._border is None:
            self._border = CellBorder()
        return self._border

    @property
    def alignment(self):
        """返回用于设置单元格对齐选项的对象"""
        if self._alignment is None:
            self._alignment = CellAlignment()
        return self._alignment

    @property
    def pattern_fill(self):
        """返回用于设置单元格图案填充的对象"""
        self._gradient_fill = None
        if self._pattern_fill is None:
            self._pattern_fill = CellPatternFill()
        return self._pattern_fill

    @property
    def gradient_fill(self):
        """返回用于设置单元格渐变填充的对象"""
        self._pattern_fill = None
        if self._gradient_fill is None:
            self._gradient_fill = CellGradientFill()
        return self._gradient_fill

    @property
    def number_format(self):
        """返回用于设置单元格数字格式的对象"""
        if self._number_format is None:
            self._number_format = CellNumberFormat()
        return self._number_format

    @property
    def protection(self):
        """返回用于设置单元格保护选项的对象"""
        if self._protection is None:
            self._protection = CellProtection()
        return self._protection

    def to_cell(self, cell, replace=True):
        """把当前样式复制到目标单元格
        :param cell: 被设置样式的单元格对象
        :param replace: 是否直接替换目标单元格的样式，是的话效率较高，但不能保留未被设置的原有样式项
        :return: None
        """
        if replace:
            self._replace_to_cell(cell)
        else:
            self._cover_to_cell(cell)

    def _cover_to_cell(self, cell):
        """把当前样式复制到目标单元格，只覆盖有设置的项，没有设置的原有的项不变
        :param cell: 被设置样式的单元格对象
        :return: None
        """
        if self._font:
            d = _handle_args(self.font_args, self._font, cell.font)
            d['family'] = cell.font.family
            cell.font = Font(**d)

        if self._border:
            d = _handle_args(self.border_args, self._border, cell.border)
            cell.border = Border(**d)

        if self._alignment:
            d = _handle_args(self.alignment_args, self._alignment, cell.alignment)
            cell.alignment = Alignment(**d)

        if self._pattern_fill:
            f = None if 'fills.GradientFill' in str(cell.fill) else cell.fill
            d = _handle_args(self.pattern_fill_args, self._pattern_fill, f)
            cell.fill = PatternFill(**d)

        elif self._gradient_fill:
            f = None if 'fills.PatternFill' in str(cell.fill) else cell.fill
            d = _handle_args(self.gradient_fill_args, self._gradient_fill, f)
            cell.fill = GradientFill(**d)

        if self._number_format and self._number_format.format != 'notSet':
            cell.number_format = self._number_format.format

        if self._protection:
            d = _handle_args(self.protection_args, self._protection, cell.protection)
            cell.protection = Protection(**d)

    def _replace_to_cell(self, cell):
        """把当前样式复制到目标单元格，覆盖原有的设置
        :param cell: 被设置样式的单元格对象
        :return: None
        """
        if self._font:
            if self._Font is None:
                d = _handle_args(self.font_args, self._font, None)
                self._Font = Font(**d)
            cell.font = self._Font

        if self._border:
            if self._Border is None:
                d = _handle_args(self.border_args, self._border, None)
                self._Border = Border(**d)
            cell.border = self._Border

        if self._alignment:
            if self._Alignment is None:
                d = _handle_args(self.alignment_args, self._alignment, None)
                self._Alignment = Alignment(**d)
            cell.alignment = self._Alignment

        if self._pattern_fill:
            if not isinstance(self._Fill, PatternFill):
                d = _handle_args(self.pattern_fill_args, self._pattern_fill, None)
                self._Fill = PatternFill(**d)
            cell.fill = self._Fill

        elif self._gradient_fill:
            if not isinstance(self._Fill, GradientFill):
                d = _handle_args(self.gradient_fill_args, self._gradient_fill, None)
                self._Fill = GradientFill(**d)
            cell.fill = self._Fill

        if self._number_format and self._number_format.format != 'notSet':
            cell.number_format = self._number_format.format

        if self._protection:
            if self._Protection is None:
                d = _handle_args(self.protection_args, self._protection, None)
                self._Protection = Protection(**d)
            cell.protection = self._Protection


def _handle_args(args, src, target=None):
    d = {}
    for arg in args:
        tmp = getattr(src, arg)
        if tmp != 'notSet':
            d[arg] = tmp
        elif target:
            d[arg] = getattr(target, arg)
    return d


class CellFont(object):
    _LINE_STYLES = ('single', 'double', 'singleAccounting', 'doubleAccounting', None)
    _SCHEMES = ('major', 'minor', None)
    _VERT_ALIGNS = ('superscript', 'subscript', 'baseline', None)

    def __init__(self):
        self.name = 'notSet'
        self.charset = 'notSet'
        self.size = 'notSet'
        self.bold = 'notSet'
        self.italic = 'notSet'
        self.strike = 'notSet'
        self.outline = 'notSet'
        self.shadow = 'notSet'
        self.condense = 'notSet'
        self.extend = 'notSet'
        self.underline = 'notSet'
        self.vertAlign = 'notSet'
        self.color = 'notSet'
        self.scheme = 'notSet'

    def set_name(self, name):
        """设置字体
        :param name: 字体名称，None表示恢复默认
        :return: None
        """
        self.name = name

    def set_charset(self, charset):
        """设置编码
        :param charset: 字体编码，int格式，None表示恢复默认
        :return: None
        """
        if not isinstance(charset, int):
            raise TypeError('charset参数只能接收int类型。')
        self.charset = charset

    def set_size(self, size):
        """设置字体大小
        :param size: 字体大小，None表示恢复默认
        :return: None
        """
        self.size = size

    def set_bold(self, on_off):
        """设置是否加粗
        :param on_off: bool表示开关，None表示恢复默认
        :return: None
        """
        self.bold = on_off

    def set_italic(self, on_off):
        """设置是否斜体
        :param on_off: bool表示开关，None表示恢复默认
        :return: None
        """
        self.italic = on_off

    def set_strike(self, on_off):
        """设置是否有删除线
        :param on_off: bool表示开关，None表示恢复默认
        :return: None
        """
        self.strike = on_off

    def set_outline(self, on_off):
        """设置outline
        :param on_off: bool表示开关，None表示恢复默认
        :return: None
        """
        self.outline = on_off

    def set_shadow(self, on_off):
        """设置是否有阴影
        :param on_off: bool表示开关，None表示恢复默认
        :return: None
        """
        self.shadow = on_off

    def set_condense(self, on_off):
        """设置condense
        :param on_off: bool表示开关，None表示恢复默认
        :return: None
        """
        self.condense = on_off

    def set_extend(self, on_off):
        """设置extend
        :param on_off: bool表示开关，None表示恢复默认
        :return: None
        """
        self.extend = on_off

    def set_color(self, color):
        """设置字体颜色
        :param color: 字体颜色，格式：'FFFFFF', '255,255,255', (255, 255, 255), Color对象均可，None表示恢复默认
        :return: None
        """
        self.color = get_color_code(color)

    def set_underline(self, option):
        """设置下划线
        :param option: 下划线类型，可选 'single', 'double', 'singleAccounting', 'doubleAccounting'，None表示恢复默认
        :return: None
        """
        if option not in self._LINE_STYLES:
            raise ValueError(f'option参数只能是{self._LINE_STYLES}其中之一。')
        self.underline = option

    def set_vertAlign(self, option):
        """设置上下标
        :param option: 可选 'superscript', 'subscript', 'baseline'，None表示恢复默认
        :return: None
        """
        if option not in self._VERT_ALIGNS:
            raise ValueError(f'option参数只能是{self._VERT_ALIGNS}其中之一。')
        self.vertAlign = option

    def set_scheme(self, option):
        """设置scheme
        :param option: 可选 'major', 'minor'，None表示恢复默认
        :return: None
        """
        if option not in self._SCHEMES:
            raise ValueError(f'option参数只能是{self._SCHEMES}其中之一。')
        self.scheme = option


class CellBorder(object):
    _LINE_STYLES = ('dashDot', 'dashDotDot', 'dashed', 'dotted', 'double', 'hair', 'medium', 'mediumDashDot',
                    'mediumDashDotDot', 'mediumDashed', 'slantDashDot', 'thick', 'thin', None)

    def __init__(self):
        self.start = 'notSet'
        self.end = 'notSet'
        self.left = 'notSet'
        self.right = 'notSet'
        self.top = 'notSet'
        self.bottom = 'notSet'
        self.diagonal = 'notSet'
        self.vertical = 'notSet'
        self.horizontal = 'notSet'
        self.horizontal = 'notSet'
        self.outline = 'notSet'
        self.diagonalUp = 'notSet'
        self.diagonalDown = 'notSet'

    def set_start(self, style, color):
        """设置start
        :param style: 线形，'dashDot','dashDotDot', 'dashed','dotted', 'double','hair', 'medium', 'mediumDashDot',
                      'mediumDashDotDot', 'mediumDashed', 'slantDashDot', 'thick', 'thin'，None表示恢复默认
        :param color: 边框颜色，格式：'FFFFFF', '255,255,255', (255, 255, 255), Color对象均可，None表示恢复默认
        :return: None
        """
        if style not in self._LINE_STYLES:
            raise ValueError(f'style参数只能是{self._LINE_STYLES}之一。')
        self.start = Side(style=style, color=get_color_code(color))

    def set_end(self, style, color):
        """设置end
        :param style: 线形，'dashDot','dashDotDot', 'dashed','dotted', 'double','hair', 'medium', 'mediumDashDot',
                      'mediumDashDotDot', 'mediumDashed', 'slantDashDot', 'thick', 'thin'，None表示恢复默认
        :param color: 边框颜色，格式：'FFFFFF', '255,255,255', (255, 255, 255), Color对象均可，None表示恢复默认
        :return: None
        """
        if style not in self._LINE_STYLES:
            raise ValueError(f'style参数只能是{self._LINE_STYLES}之一。')
        self.end = Side(style=style, color=get_color_code(color))

    def set_left(self, style, color):
        """设置左边框
        :param style: 线形，'dashDot','dashDotDot', 'dashed','dotted', 'double','hair', 'medium', 'mediumDashDot',
                      'mediumDashDotDot', 'mediumDashed', 'slantDashDot', 'thick', 'thin'，None表示恢复默认
        :param color: 边框颜色，格式：'FFFFFF', '255,255,255', (255, 255, 255), Color对象均可，None表示恢复默认
        :return: None
        """
        if style not in self._LINE_STYLES:
            raise ValueError(f'style参数只能是{self._LINE_STYLES}之一。')
        self.left = Side(style=style, color=get_color_code(color))

    def set_right(self, style, color):
        """设置右边框
        :param style: 线形，'dashDot','dashDotDot', 'dashed','dotted', 'double','hair', 'medium', 'mediumDashDot',
                      'mediumDashDotDot', 'mediumDashed', 'slantDashDot', 'thick', 'thin'，None表示恢复默认
        :param color: 边框颜色，格式：'FFFFFF', '255,255,255', (255, 255, 255), Color对象均可，None表示恢复默认
        :return: None
        """
        if style not in self._LINE_STYLES:
            raise ValueError(f'style参数只能是{self._LINE_STYLES}之一。')
        self.right = Side(style=style, color=get_color_code(color))

    def set_top(self, style, color):
        """设置上边框
        :param style: 线形，'dashDot','dashDotDot', 'dashed','dotted', 'double','hair', 'medium', 'mediumDashDot',
                      'mediumDashDotDot', 'mediumDashed', 'slantDashDot', 'thick', 'thin'，None表示恢复默认
        :param color: 边框颜色，格式：'FFFFFF', '255,255,255', (255, 255, 255), Color对象均可，None表示恢复默认
        :return: None
        """
        if style not in self._LINE_STYLES:
            raise ValueError(f'style参数只能是{self._LINE_STYLES}之一。')
        self.top = Side(style=style, color=get_color_code(color))

    def set_bottom(self, style, color):
        """设置下边框
        :param style: 线形，'dashDot','dashDotDot', 'dashed','dotted', 'double','hair', 'medium', 'mediumDashDot',
                      'mediumDashDotDot', 'mediumDashed', 'slantDashDot', 'thick', 'thin'，None表示恢复默认
        :param color: 边框颜色，格式：'FFFFFF', '255,255,255', (255, 255, 255), Color对象均可，None表示恢复默认
        :return: None
        """
        if style not in self._LINE_STYLES:
            raise ValueError(f'style参数只能是{self._LINE_STYLES}之一。')
        self.bottom = Side(style=style, color=get_color_code(color))

    def set_diagonal(self, style, color):
        """设置对角线
        :param style: 线形，'dashDot','dashDotDot', 'dashed','dotted', 'double','hair', 'medium', 'mediumDashDot',
                      'mediumDashDotDot', 'mediumDashed', 'slantDashDot', 'thick', 'thin'，None表示恢复默认
        :param color: 边框颜色，格式：'FFFFFF', '255,255,255', (255, 255, 255), Color对象均可，None表示恢复默认
        :return: None
        """
        if style not in self._LINE_STYLES:
            raise ValueError(f'style参数只能是{self._LINE_STYLES}之一。')
        self.diagonal = Side(style=style, color=get_color_code(color))

    def set_vertical(self, style, color):
        """设置垂直中线
        :param style: 线形，'dashDot','dashDotDot', 'dashed','dotted', 'double','hair', 'medium', 'mediumDashDot',
                      'mediumDashDotDot', 'mediumDashed', 'slantDashDot', 'thick', 'thin'，None表示恢复默认
        :param color: 边框颜色，格式：'FFFFFF', '255,255,255', (255, 255, 255), Color对象均可，None表示恢复默认
        :return: None
        """
        if style not in self._LINE_STYLES:
            raise ValueError(f'style参数只能是{self._LINE_STYLES}之一。')
        self.vertical = Side(style=style, color=get_color_code(color))

    def set_horizontal(self, style, color):
        """设置水平中线
        :param style: 线形，'dashDot','dashDotDot', 'dashed','dotted', 'double','hair', 'medium', 'mediumDashDot',
                      'mediumDashDotDot', 'mediumDashed', 'slantDashDot', 'thick', 'thin'，None表示恢复默认
        :param color: 边框颜色，格式：'FFFFFF', '255,255,255', (255, 255, 255), Color对象均可，None表示恢复默认
        :return: None
        """
        if style not in self._LINE_STYLES:
            raise ValueError(f'style参数只能是{self._LINE_STYLES}之一。')
        self.horizontal = Side(style=style, color=get_color_code(color))

    def set_outline(self, on_off):
        """
        :param on_off: bool表示开关
        :return: None
        """
        self.outline = on_off

    def set_diagonalDown(self, on_off):
        """
        :param on_off: bool表示开关
        :return: None
        """
        self.diagonalDown = on_off

    def set_diagonalUp(self, on_off):
        """
        :param on_off: bool表示开关
        :return: None
        """
        self.diagonalUp = on_off


class CellAlignment(object):
    _horizontal_alignments = ('general', 'left', 'center', 'right', 'fill', 'justify', 'centerContinuous',
                              'distributed', None)
    _vertical_alignments = ('top', 'center', 'bottom', 'justify', 'distributed', None)

    def __init__(self):
        self.horizontal = 'notSet'
        self.vertical = 'notSet'
        self.indent = 'notSet'
        self.relativeIndent = 'notSet'
        self.justifyLastLine = 'notSet'
        self.readingOrder = 'notSet'
        self.textRotation = 'notSet'
        self.wrapText = 'notSet'
        self.shrinkToFit = 'notSet'

    def set_horizontal(self, horizontal):
        """设置水平位置
        :param horizontal: 可选：'general', 'left', 'center', 'right', 'fill', 'justify', 'centerContinuous',
                                'distributed'，None表示恢复默认
        :return: None
        """
        if horizontal not in self._horizontal_alignments:
            raise ValueError(f'horizontal参数必须是{self._horizontal_alignments}其中之一。')
        self.horizontal = horizontal

    def set_vertical(self, vertical):
        """设置垂直位置
        :param vertical: 可选：'top', 'center', 'bottom', 'justify', 'distributed'，None表示恢复默认
        :return: None
        """
        if vertical not in self._vertical_alignments:
            raise ValueError(f'horizontal参数必须是{self._vertical_alignments}其中之一。')
        self.vertical = vertical

    def set_indent(self, indent):
        """设置缩进
        :param indent: 缩进数值，0到255
        :return: None
        """
        if not (isinstance(indent, int) and 0 <= indent <= 255):
            raise ValueError('value参数必须在0到255之间。')
        self.indent = indent

    def set_relativeIndent(self, indent):
        """设置相对缩进
        :param indent: 缩进数值，-255到255
        :return: None
        """
        if not (isinstance(indent, int) and -255 <= indent <= 255):
            raise ValueError('value参数必须在-255到255之间。')
        self.relativeIndent = indent

    def set_justifyLastLine(self, on_off):
        """设置justifyLastLine
        :param on_off: bool表示开或关，None表示恢复默认
        :return: None
        """
        self.justifyLastLine = on_off

    def set_readingOrder(self, value):
        """设置readingOrder
        :param value: 不小于0的数字
        :return: None
        """
        if not (isinstance(value, int) and value >= 0):
            raise ValueError('value参数必须不小于0。')
        self.readingOrder = value

    def set_textRotation(self, value):
        """设置文本旋转角度
        :param value: 0-180或255
        :return: None
        """
        if not (0 <= value <= 180 or value == 255):
            raise ValueError('value必须在0到180之间。')
        self.textRotation = value

    def set_wrapText(self, on_off):
        """设置wrapText
        :param on_off: bool表示开或关，None表示恢复默认
        :return: None
        """
        self.wrapText = on_off

    def set_shrinkToFit(self, on_off):
        """设置shrinkToFit
        :param on_off: bool表示开或关，None表示恢复默认
        :return: None
        """
        self.shrinkToFit = on_off


class CellGradientFill(object):
    def __init__(self):
        self.type = 'notSet'
        self.degree = 'notSet'
        self.left = 'notSet'
        self.right = 'notSet'
        self.top = 'notSet'
        self.bottom = 'notSet'
        self.stop = 'notSet'

    def set_type(self, name):
        """设置类型
        :param name: 可选：'linear', 'path'
        :return: None
        """
        if name not in ('linear', 'path'):
            raise ValueError("name参数只能是 'linear', 'path' 之一。")
        self.type = name

    def set_degree(self, value):
        """设置程度
        :param value: 数值
        :return: None
        """
        self.degree = value

    def set_left(self, value):
        """设置left
        :param value: 数值
        :return: None
        """
        self.left = value

    def set_right(self, value):
        """设置right
        :param value: 数值
        :return: None
        """
        self.right = value

    def set_top(self, value):
        """设置top
        :param value: 数值
        :return: None
        """
        self.top = value

    def set_bottom(self, value):
        """设置bottom
        :param value: 数值
        :return: None
        """
        self.bottom = value

    def set_stop(self, values):
        """设置stop
        :param values: 数值
        :return: None
        """
        self.stop = values


class CellPatternFill(object):
    _FILES = ('none', 'solid', 'darkDown', 'darkGray', 'darkGrid', 'darkHorizontal', 'darkTrellis', 'darkUp',
              'darkVertical', 'gray0625', 'gray125', 'lightDown', 'lightGray', 'lightGrid', 'lightHorizontal',
              'lightTrellis', 'lightUp', 'lightVertical', 'mediumGray', None)

    def __init__(self):
        self.patternType = 'notSet'
        self.fgColor = 'notSet'
        self.bgColor = 'notSet'

    def set_patternType(self, name):
        """设置类型
        :param name: 可选：'none', 'solid', 'darkDown', 'darkGray', 'darkGrid', 'darkHorizontal', 'darkTrellis',
                          'darkUp', 'darkVertical', 'gray0625', 'gray125', 'lightDown', 'lightGray', 'lightGrid',
                          'lightHorizontal', 'lightTrellis', 'lightUp', 'lightVertical', 'mediumGray'，None为恢复默认
        :return: None
        """
        if name not in self._FILES:
            raise ValueError(f'name参数只能是{self._FILES}其中之一。')
        self.patternType = name

    def set_fgColor(self, color):
        """设置前景色
        :param color: 颜色，格式：'FFFFFF', '255,255,255', (255, 255, 255), Color对象均可，None表示恢复默认
        :return: None
        """
        self.fgColor = get_color_code(color)

    def set_bgColor(self, color):
        """设置背景色
        :param color: 颜色，格式：'FFFFFF', '255,255,255', (255, 255, 255), Color对象均可，None表示恢复默认
        :return: None
        """
        self.bgColor = get_color_code(color)


class CellNumberFormat(object):
    def __init__(self):
        self.format = 'notSet'

    def set_format(self, string):
        """设置数字格式
        :param string: 格式字符串，为None时恢复默认，格式很多具体在`openpyxl.numbers`查看
        :return: None
        """
        if string is None:
            string = 'General'
        self.format = string


class CellProtection(object):
    def __init__(self):
        self.hidden = 'notSet'
        self.locked = 'notSet'

    def set_hidden(self, on_off):
        """设置是否隐藏
        :param on_off: bool表示开关
        :return: None
        """
        self.hidden = on_off

    def set_locked(self, on_off):
        """设置是否锁定
        :param on_off: bool表示开关
        :return: None
        """
        self.locked = on_off


class CellStyleCopier(object):
    def __init__(self, from_cell):
        """
        :param from_cell: 被复制单元格对象
        """
        self._style = copy(from_cell._style)
        self._font = copy(from_cell.font)
        self._border = copy(from_cell.border)
        self._fill = copy(from_cell.fill)
        self._number_format = copy(from_cell.number_format)
        self._protection = copy(from_cell.protection)
        self._alignment = copy(from_cell.alignment)

    def to_cell(self, cell):
        """把当前样式复制到目标单元格
        :param cell: 被设置样式的单元格对象
        :return: None
        """
        cell._style = self._style
        cell.alignment = self._alignment
        cell.font = self._font
        cell.border = self._border
        cell.fill = self._fill
        cell.number_format = self._number_format
        cell.protection = self._protection


def get_color_code(color):
    """将颜色拼音转为代码
    :param color: 颜色名称或代码字符串
    :return: 颜色代码
    """
    if color is None:
        return '000000'
    if isinstance(color, Color):
        return color
    __COLORS__ = {
        'white': 'FFFFFF',
        'black': '000000',
        'red': 'FF0000',
        'green': '7FB80E',
        'blue': '009AD6',
        'purple': '8552A1',
        'yellow': 'FFFF00',
        'orange': 'F58220'
    }
    color = str(color)
    if ',' in color:
        color = color.replace(' ', '').lstrip('(').rstrip(')')
        RGB = color.split(',')
        color = ''
        for i in RGB:
            num = int(i)
            color += str(hex(num))[-2:].replace('x', '0').upper()
        return color

    return __COLORS__.get(color, color).lstrip('#')


class NoneStyle(object):
    _instance_lock = Lock()

    def __init__(self):
        self._font = Font()
        self._border = Border()
        self._alignment = Alignment()
        self._fill = PatternFill()
        self._number_format = 'General'
        self._protection = Protection()

    def __new__(cls, *args, **kwargs):
        if not hasattr(NoneStyle, "_instance"):
            with NoneStyle._instance_lock:
                if not hasattr(NoneStyle, "_instance"):
                    NoneStyle._instance = object.__new__(cls)
        return NoneStyle._instance

    def to_cell(self, cell, replace=True):
        cell.font = self._font
        cell.border = self._border
        cell.alignment = self._alignment
        cell.fill = self._fill
        cell.protection = self._protection
        cell.number_format = 'General'
