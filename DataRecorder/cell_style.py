# -*- coding:utf-8 -*-
from openpyxl.styles import Alignment, Font, Side, Border, Protection, GradientFill, PatternFill


class CellStyle(object):
    font_args = ('name', 'size', 'charset', 'underline', 'color', 'scheme', 'vertAlign',
                 'bold', 'italic', 'strike', 'outline', 'shadow', 'condense', 'extend')
    border_args = ('start', 'end', 'left', 'right', 'top', 'bottom', 'diagonal', 'vertical', 'horizontal',
                   'horizontal', 'outline', 'diagonalUp', 'diagonalDown')
    alignment_args = ('horizontal', 'vertical', 'indent', 'relativeIndent', 'justifyLastLine', 'readingOrder',
                      'textRotation', 'wrapText', 'shrinkToFit')
    protection_args = ('locked', 'hidden')
    gradient_fill_args = ('type', 'degree', 'left', 'right', 'top', 'bottom', 'stop')
    pattern_fill_args = ('patternType', 'fgColor', 'bgColor')

    def __init__(self):
        self._style = None
        self._font = None
        self._border = None
        self._alignment = None
        self._pattern_fill = None
        self._gradient_fill = None
        self._number_format = None
        self._protection = None

    @property
    def font(self):
        if self._font is None:
            self._font = CellFont()
        return self._font

    @property
    def border(self):
        if self._border is None:
            self._border = CellBorder()
        return self._border

    @property
    def alignment(self):
        if self._alignment is None:
            self._alignment = CellAlignment()
        return self._alignment

    @property
    def pattern_fill(self):
        self._gradient_fill = None
        if self._pattern_fill is None:
            self._pattern_fill = CellPatternFill()
        return self._pattern_fill

    @property
    def gradient_fill(self):
        self._pattern_fill = None
        if self._gradient_fill is None:
            self._gradient_fill = CellGradientFill()
        return self._gradient_fill

    @property
    def CellNumFormat(self):
        if self._number_format is None:
            self._number_format = CellAlignment()
        return self._number_format

    @property
    def protection(self):
        if self._protection is None:
            self._protection = CellProtection()
        return self._protection

    def to_cell(self, cell):
        """把当前样式复制到目标单元格"""
        if self._font:
            d = _handle_args(self.font_args, self._font, cell.font)
            d['family'] = cell.font.family
            cell.font = Font(**d)

        if self._border:
            d = _handle_args(self.border_args, self._border, cell.border)
            cell.border = Border(**d)

        if self._alignment:
            d = _handle_args(self.alignment_args, self._alignment, cell.alignment)
            cell.alignment = Alignment(**d)

        if self._pattern_fill:
            f = None if 'fills.GradientFill' in str(cell.fill) else cell.fill
            d = _handle_args(self.pattern_fill_args, self._pattern_fill, f)
            cell.fill = PatternFill(**d)

        elif self._gradient_fill:
            f = None if 'fills.PatternFill' in str(cell.fill) else cell.fill
            d = _handle_args(self.gradient_fill_args, self._gradient_fill, f)
            cell.fill = GradientFill(**d)

        if self._number_format:
            cell.number_format = self._number_format

        if self._protection:
            d = _handle_args(self.protection_args, self._protection, cell.protection)
            cell.protection = Protection(**d)


def _handle_args(args, src, target=None):
    d = {}
    for arg in args:
        tmp = getattr(src, arg)
        if tmp != 'notSet':
            d[arg] = tmp
        elif target:
            d[arg] = getattr(target, arg)
    return d


class CellFont(object):
    LINE_STYLES = ('single', 'double', 'singleAccounting', 'doubleAccounting', None)
    SCHEMES = ('major', 'minor', None)
    VERT_ALIGNS = ('superscript', 'subscript', 'baseline', None)

    def __init__(self):
        self.name = 'notSet'
        self.charset = 'notSet'
        self.size = 'notSet'
        self.bold = 'notSet'
        self.italic = 'notSet'
        self.strike = 'notSet'
        self.outline = 'notSet'
        self.shadow = 'notSet'
        self.condense = 'notSet'
        self.extend = 'notSet'
        self.underline = 'notSet'
        self.vertAlign = 'notSet'
        self.color = 'notSet'
        self.scheme = 'notSet'

    def set_name(self, name):
        """设置字体
        :param name: 字体名称，None表示恢复默认
        :return: None
        """
        self.name = name

    def set_charset(self, charset):
        """设置编码
        :param charset: 字体编码，int格式，None表示恢复默认
        :return: None
        """
        if not isinstance(charset, int):
            raise TypeError('charset参数只能接收int类型。')
        self.charset = charset

    def set_size(self, size):
        """设置字体大小
        :param size: 字体大小，None表示恢复默认
        :return: None
        """
        self.size = size

    def set_bold(self, on_off):
        """设置是否加粗
        :param on_off: bool表示开关，None表示恢复默认
        :return: None
        """
        self.bold = on_off

    def set_italic(self, on_off):
        """设置是否斜体
        :param on_off: bool表示开关，None表示恢复默认
        :return: None
        """
        self.italic = on_off

    def set_strike(self, on_off):
        """设置是否有删除线
        :param on_off: bool表示开关，None表示恢复默认
        :return: None
        """
        self.strike = on_off

    def set_outline(self, on_off):
        """设置outline
        :param on_off: bool表示开关，None表示恢复默认
        :return: None
        """
        self.outline = on_off

    def set_shadow(self, on_off):
        """设置是否有阴影
        :param on_off: bool表示开关，None表示恢复默认
        :return: None
        """
        self.shadow = on_off

    def set_condense(self, on_off):
        """设置condense
        :param on_off: bool表示开关，None表示恢复默认
        :return: None
        """
        self.condense = on_off

    def set_extend(self, on_off):
        """设置extend
        :param on_off: bool表示开关，None表示恢复默认
        :return: None
        """
        self.extend = on_off

    def set_color(self, color):
        """设置字体颜色
        :param color: 字体演示字符串，如`FFF000`，None表示恢复默认
        :return: None
        """
        self.color = color

    def set_underline(self, option):
        """设置下划线
        :param option: 下划线类型，可选 'single', 'double', 'singleAccounting', 'doubleAccounting'，None表示恢复默认
        :return: None
        """
        if option not in self.LINE_STYLES:
            raise ValueError(f'option参数只能是{self.LINE_STYLES}其中之一。')
        self.underline = option

    def set_vertAlign(self, option):
        """设置上下标
        :param option: 可选 'superscript', 'subscript', 'baseline'，None表示恢复默认
        :return: None
        """
        if option not in self.VERT_ALIGNS:
            raise ValueError(f'option参数只能是{self.VERT_ALIGNS}其中之一。')
        self.vertAlign = option

    def set_scheme(self, option):
        """设置scheme
        :param option: 可选 'major', 'minor'，None表示恢复默认
        :return: None
        """
        if option not in self.SCHEMES:
            raise ValueError(f'option参数只能是{self.SCHEMES}其中之一。')
        self.scheme = option


class CellBorder(object):
    LINE_STYLES = ('dashDot', 'dashDotDot', 'dashed', 'dotted', 'double', 'hair', 'medium', 'mediumDashDot',
                   'mediumDashDotDot', 'mediumDashed', 'slantDashDot', 'thick', 'thin', None)

    def __init__(self):
        self.start = 'notSet'
        self.end = 'notSet'
        self.left = 'notSet'
        self.right = 'notSet'
        self.top = 'notSet'
        self.bottom = 'notSet'
        self.diagonal = 'notSet'
        self.vertical = 'notSet'
        self.horizontal = 'notSet'
        self.horizontal = 'notSet'
        self.outline = 'notSet'
        self.diagonalUp = 'notSet'
        self.diagonalDown = 'notSet'

    def set_start(self, style, color):
        """设置start
        :param style: 线形，'dashDot','dashDotDot', 'dashed','dotted', 'double','hair', 'medium', 'mediumDashDot',
                      'mediumDashDotDot', 'mediumDashed', 'slantDashDot', 'thick', 'thin'，None表示恢复默认
        :param color: 线条颜色
        :return: None
        """
        if style not in self.LINE_STYLES:
            raise ValueError(f'style参数只能是{self.LINE_STYLES}之一。')
        self.start = Side(style=style, color=color)

    def set_end(self, style, color):
        """设置end
        :param style: 线形，'dashDot','dashDotDot', 'dashed','dotted', 'double','hair', 'medium', 'mediumDashDot',
                      'mediumDashDotDot', 'mediumDashed', 'slantDashDot', 'thick', 'thin'，None表示恢复默认
        :param color: 线条颜色
        :return: None
        """
        if style not in self.LINE_STYLES:
            raise ValueError(f'style参数只能是{self.LINE_STYLES}之一。')
        self.end = Side(style=style, color=color)

    def set_left(self, style, color):
        """设置左边框
        :param style: 线形，'dashDot','dashDotDot', 'dashed','dotted', 'double','hair', 'medium', 'mediumDashDot',
                      'mediumDashDotDot', 'mediumDashed', 'slantDashDot', 'thick', 'thin'，None表示恢复默认
        :param color: 线条颜色
        :return: None
        """
        if style not in self.LINE_STYLES:
            raise ValueError(f'style参数只能是{self.LINE_STYLES}之一。')
        self.left = Side(style=style, color=color)

    def set_right(self, style, color):
        """设置右边框
        :param style: 线形，'dashDot','dashDotDot', 'dashed','dotted', 'double','hair', 'medium', 'mediumDashDot',
                      'mediumDashDotDot', 'mediumDashed', 'slantDashDot', 'thick', 'thin'，None表示恢复默认
        :param color: 线条颜色
        :return: None
        """
        if style not in self.LINE_STYLES:
            raise ValueError(f'style参数只能是{self.LINE_STYLES}之一。')
        self.right = Side(style=style, color=color)

    def set_top(self, style, color):
        """设置上边框
        :param style: 线形，'dashDot','dashDotDot', 'dashed','dotted', 'double','hair', 'medium', 'mediumDashDot',
                      'mediumDashDotDot', 'mediumDashed', 'slantDashDot', 'thick', 'thin'，None表示恢复默认
        :param color: 线条颜色
        :return: None
        """
        if style not in self.LINE_STYLES:
            raise ValueError(f'style参数只能是{self.LINE_STYLES}之一。')
        self.top = Side(style=style, color=color)

    def set_bottom(self, style, color):
        """设置下边框
        :param style: 线形，'dashDot','dashDotDot', 'dashed','dotted', 'double','hair', 'medium', 'mediumDashDot',
                      'mediumDashDotDot', 'mediumDashed', 'slantDashDot', 'thick', 'thin'，None表示恢复默认
        :param color: 线条颜色
        :return: None
        """
        if style not in self.LINE_STYLES:
            raise ValueError(f'style参数只能是{self.LINE_STYLES}之一。')
        self.bottom = Side(style=style, color=color)

    def set_diagonal(self, style, color):
        """设置对角线
        :param style: 线形，'dashDot','dashDotDot', 'dashed','dotted', 'double','hair', 'medium', 'mediumDashDot',
                      'mediumDashDotDot', 'mediumDashed', 'slantDashDot', 'thick', 'thin'，None表示恢复默认
        :param color: 线条颜色
        :return: None
        """
        if style not in self.LINE_STYLES:
            raise ValueError(f'style参数只能是{self.LINE_STYLES}之一。')
        self.diagonal = Side(style=style, color=color)

    def set_vertical(self, style, color):
        """设置垂直中线
        :param style: 线形，'dashDot','dashDotDot', 'dashed','dotted', 'double','hair', 'medium', 'mediumDashDot',
                      'mediumDashDotDot', 'mediumDashed', 'slantDashDot', 'thick', 'thin'，None表示恢复默认
        :param color: 线条颜色
        :return: None
        """
        if style not in self.LINE_STYLES:
            raise ValueError(f'style参数只能是{self.LINE_STYLES}之一。')
        self.vertical = Side(style=style, color=color)

    def set_horizontal(self, style, color):
        """设置水平中线
        :param style: 线形，'dashDot','dashDotDot', 'dashed','dotted', 'double','hair', 'medium', 'mediumDashDot',
                      'mediumDashDotDot', 'mediumDashed', 'slantDashDot', 'thick', 'thin'，None表示恢复默认
        :param color: 线条颜色
        :return: None
        """
        if style not in self.LINE_STYLES:
            raise ValueError(f'style参数只能是{self.LINE_STYLES}之一。')
        self.horizontal = Side(style=style, color=color)

    def set_outline(self, on_off):
        """
        :param on_off: bool表示开关
        :return: None
        """
        self.outline = on_off

    def set_diagonalDown(self, on_off):
        """
        :param on_off: bool表示开关
        :return: None
        """
        self.diagonalDown = on_off

    def set_diagonalUp(self, on_off):
        """
        :param on_off: bool表示开关
        :return: None
        """
        self.diagonalUp = on_off


class CellAlignment(object):
    horizontal_alignments = ('general', 'left', 'center', 'right', 'fill', 'justify', 'centerContinuous',
                             'distributed', None)
    vertical_alignments = ('top', 'center', 'bottom', 'justify', 'distributed', None)

    def __init__(self):
        self.horizontal = 'notSet'
        self.vertical = 'notSet'
        self.indent = 'notSet'
        self.relativeIndent = 'notSet'
        self.justifyLastLine = 'notSet'
        self.readingOrder = 'notSet'
        self.textRotation = 'notSet'
        self.wrapText = 'notSet'
        self.shrinkToFit = 'notSet'

    def set_horizontal(self, horizontal):
        """设置水平位置
        :param horizontal: 可选："general", "left", "center", "right", "fill", "justify", "centerContinuous",
                                "distributed"，None表示恢复默认
        :return: None
        """
        if horizontal not in self.horizontal_alignments:
            raise ValueError(f'horizontal参数必须是{self.horizontal_alignments}其中之一。')
        self.horizontal = horizontal

    def set_vertical(self, vertical):
        """设置垂直位置
        :param vertical: 可选："top", "center", "bottom", "justify", "distributed"，None表示恢复默认
        :return: None
        """
        if vertical not in self.vertical_alignments:
            raise ValueError(f'horizontal参数必须是{self.vertical_alignments}其中之一。')
        self.vertical = vertical

    def set_indent(self, indent):
        """设置缩进
        :param indent: 缩进数值，0到255
        :return: None
        """
        if not (isinstance(indent, int) and 0 <= indent <= 255):
            raise ValueError('value参数必须在0到255之间。')
        self.indent = indent

    def set_relativeIndent(self, indent):
        """设置相对缩进
        :param indent: 缩进数值，-255到255
        :return: None
        """
        if not (isinstance(indent, int) and -255 <= indent <= 255):
            raise ValueError('value参数必须在-255到255之间。')
        self.relativeIndent = indent

    def set_justifyLastLine(self, on_off):
        """设置justifyLastLine
        :param on_off: bool表示开或关，None表示恢复默认
        :return: None
        """
        self.justifyLastLine = on_off

    def set_readingOrder(self, value):
        """设置readingOrder
        :param value: 不小于0的数字
        :return: None
        """
        if not (isinstance(value, int) and value >= 0):
            raise ValueError('value参数必须不小于0。')
        self.readingOrder = value

    def set_textRotation(self, value):
        """设置文本旋转角度
        :param value: 0-180
        :return: None
        """
        if not (0 <= value <= 180 or value == 255):
            raise ValueError('value必须在0到180之间。')
        self.textRotation = value

    def set_wrapText(self, on_off):
        """设置wrapText
        :param on_off: bool表示开或关，None表示恢复默认
        :return: None
        """
        self.wrapText = on_off

    def set_shrinkToFit(self, on_off):
        """设置shrinkToFit
        :param on_off: bool表示开或关，None表示恢复默认
        :return: None
        """
        self.shrinkToFit = on_off


class CellGradientFill(object):
    def __init__(self):
        self.type = 'notSet'
        self.degree = 'notSet'
        self.left = 'notSet'
        self.right = 'notSet'
        self.top = 'notSet'
        self.bottom = 'notSet'
        self.stop = 'notSet'

    def set_type(self, name):
        """设置类型
        :param name: 可选：'linear', 'path'
        :return: None
        """
        if name not in ('linear', 'path'):
            raise ValueError("name参数只能是 'linear', 'path' 之一。")
        self.type = name

    def set_degree(self, value):
        """设置程度
        :param value: 数值
        :return: None
        """
        self.degree = value

    def set_left(self, value):
        """设置left
        :param value: 数值
        :return: None
        """
        self.left = value

    def set_right(self, value):
        """设置right
        :param value: 数值
        :return: None
        """
        self.right = value

    def set_top(self, value):
        """设置top
        :param value: 数值
        :return: None
        """
        self.top = value

    def set_bottom(self, value):
        """设置bottom
        :param value: 数值
        :return: None
        """
        self.bottom = value

    def set_stop(self, value):
        """设置stop
        :param value: 数值
        :return: None
        """
        self.stop = value


class CellPatternFill(object):
    FILES = ('none', 'solid', 'darkDown', 'darkGray', 'darkGrid', 'darkHorizontal', 'darkTrellis', 'darkUp',
             'darkVertical', 'gray0625', 'gray125', 'lightDown', 'lightGray', 'lightGrid', 'lightHorizontal',
             'lightTrellis', 'lightUp', 'lightVertical', 'mediumGray', None)

    def __init__(self):
        self.patternType = 'notSet'
        self.fgColor = 'notSet'
        self.bgColor = 'notSet'

    def set_patternType(self, name):
        """设置类型
        :param name: 可选：'none', 'solid', 'darkDown', 'darkGray', 'darkGrid', 'darkHorizontal', 'darkTrellis',
                          'darkUp', 'darkVertical', 'gray0625', 'gray125', 'lightDown', 'lightGray', 'lightGrid',
                          'lightHorizontal', 'lightTrellis', 'lightUp', 'lightVertical', 'mediumGray'，None为恢复默认
        :return: None
        """
        if name not in self.FILES:
            raise ValueError(f'name参数只能是{self.FILES}其中之一。')
        self.patternType = name

    def set_fgColor(self, color):
        """设置前景色
        :param color: 颜色字符串，None为恢复默认
        :return: None
        """
        self.fgColor = color

    def set_bgColor(self, color):
        """设置背景色
        :param color: 颜色字符串，None为恢复默认
        :return: None
        """
        self.bgColor = color


class CellNumberFormat(object):
    def __init__(self):
        pass


class CellProtection(object):
    def __init__(self):
        self.hidden = 'notSet'
        self.locked = 'notSet'

    def set_hidden(self, on_off):
        """设置是否隐藏
        :param on_off: bool表示开关
        :return: None
        """
        self.hidden = on_off

    def set_locked(self, on_off):
        """设置是否锁定
        :param on_off: bool表示开关
        :return: None
        """
        self.locked = on_off
