# -*- coding:utf-8 -*-
from pathlib import Path
from time import sleep
from typing import Union

from openpyxl import Workbook, load_workbook

from .base import BaseRecorder
from .style.cell_style import CellStyleCopier
from .setter import RecorderSetter
from .tools import ok_list, data_to_list_or_dict, process_content


class Recorder(BaseRecorder):
    SUPPORTS = ('csv', 'xlsx', 'json', 'txt')

    def __init__(self, path=None, cache_size=None):
        """用于缓存并记录数据，可在达到一定数量时自动记录，以降低文件读写次数，减少开销
        :param path: 保存的文件路径
        :param cache_size: 每接收多少条记录写入文件，0为不自动写入
        """
        super().__init__(path=path, cache_size=cache_size)
        self._delimiter = ','  # csv文件分隔符
        self._quote_char = '"'  # csv文件引用符
        self._follow_styles = False
        self._col_height = None
        self._style = None
        self._fit_head = False

    @property
    def set(self):
        """返回用于设置属性的对象"""
        if self._setter is None:
            self._setter = RecorderSetter(self)
        return self._setter

    @property
    def delimiter(self):
        """返回csv文件分隔符"""
        return self._delimiter

    @property
    def quote_char(self):
        """返回csv文件引用符"""
        return self._quote_char

    def add_data(self, data, table=None):
        """添加数据，可一次添加多条数据
        :param data: 插入的数据，任意格式
        :param table: 要写入的数据表，仅支持xlsx格式。为None表示用set.table()方法设置的值，为bool表示活动的表格
        :return: None
        """
        while self._pause_add:  # 等待其它线程写入结束
            sleep(.1)

        if not isinstance(data, (list, tuple, dict)):
            data = (data,)

        if not data:
            data = ([],)
            self._data_count += 1

        # 一维数组
        elif isinstance(data, dict) or (
                isinstance(data, (list, tuple)) and not isinstance(data[0], (list, tuple, dict))):
            data = [data_to_list_or_dict(self, data)]
            self._data_count += 1

        else:  # 二维数组
            data = [data_to_list_or_dict(self, d) for d in data]
            self._data_count += len(data)

        if self._type != 'xlsx':
            self._data.extend(data)

        else:
            if table is None:
                table = self._table
            elif isinstance(table, bool):
                table = None

            self._data.setdefault(table, []).extend(data)

        if 0 < self.cache_size <= self._data_count:
            self.record()

    def _record(self):
        """记录数据"""
        if self.type == 'csv':
            self._to_csv()
        elif self.type == 'xlsx':
            self._to_xlsx()
        elif self.type == 'json':
            self._to_json()
        elif self.type == 'txt':
            self._to_txt()

    def _to_xlsx(self):
        """记录数据到xlsx文件"""
        if Path(self.path).exists():
            new_file = False
            wb = load_workbook(self.path)

        else:
            new_file = True
            if self._col_height or self._follow_styles or self._style or len(self._data) > 1:
                wb = Workbook(write_only=False)
            else:
                wb = Workbook(write_only=True)
                wb.create_sheet('Sheet1')

        tables = [i.title for i in wb.worksheets]
        for table, data in self._data.items():
            _row_styles = None
            _col_height = None
            new_sheet = False

            if table is None:
                ws = wb.active

            elif table in tables:
                ws = wb[table]

            elif new_file:
                ws = wb.active
                tables.remove(ws.title)
                ws.title = table
                tables.append(table)
                new_sheet = True

            else:
                ws = wb.create_sheet(title=table)
                tables.append(table)
                new_sheet = True

            if new_file or new_sheet:
                new_file = False
                title = _get_title(data[0], self._before, self._after)
                if title is not None:
                    ws.append(ok_list(title, True))

            elif self._follow_styles:
                row_num = ws.max_row
                _row_styles = [CellStyleCopier(i) for i in ws[row_num]]
                _col_height = ws.row_dimensions[row_num].height
                title = [r.value for r in ws[1]]

            else:
                title = [r.value for r in ws[1]]

            if self._fit_head and not self._head.get(ws.title, None) and any(title):
                self._head[ws.title] = title

            # ====================================

            if self._fit_head and self._head.get(ws.title, None):
                for i in data:
                    if isinstance(i, dict):
                        i = [i.get(h, None) for h in self._head[ws.title]]

                    ws.append(ok_list(i, True))
                    _set_style(_col_height, _row_styles, ws, self)

            else:
                for i in data:
                    ws.append(ok_list(i, True))
                    _set_style(_col_height, _row_styles, ws, self)

        wb.save(self.path)
        wb.close()

    def _to_csv(self):
        """记录数据到csv文件"""
        if Path(self.path).exists():
            title = None
            if self._fit_head and not self._head:
                from csv import reader
                with open(self.path, 'r', newline='', encoding=self.encoding) as f:
                    u = reader(f, delimiter=self.delimiter, quotechar=self.quote_char)
                    try:
                        self._head = next(u)
                    except StopIteration:
                        pass

        else:
            title = _get_title(self._data[0], self._before, self._after)
            if self._fit_head:
                self._head = title

        with open(self.path, 'a+', newline='', encoding=self.encoding) as f:
            from csv import writer
            csv_write = writer(f, delimiter=self.delimiter, quotechar=self.quote_char)
            if title:
                csv_write.writerow(ok_list(title))

            if self._fit_head and self._head:
                for i in self._data:
                    d = data_to_list_or_dict(self, i)
                    if isinstance(d, dict):
                        d = [d.get(h, None) for h in self._head]

                    csv_write.writerow(ok_list(d))

            else:
                for i in self._data:
                    csv_write.writerow(ok_list(i))

    def _to_txt(self):
        """记录数据到txt文件"""
        with open(self.path, 'a+', encoding=self.encoding) as f:
            all_data = [' '.join(ok_list(i, as_str=True)) for i in self._data]
            f.write('\n'.join(all_data) + '\n')

    def _to_json(self):
        """记录数据到json文件"""
        from json import load, dump
        if Path(self.path).exists():
            with open(self.path, 'r', encoding=self.encoding) as f:
                json_data = load(f)

            for i in self._data:
                if isinstance(i, dict):
                    for d in i:
                        i[d] = process_content(i[d])
                    json_data.append(i)
                else:
                    json_data.append([process_content(d) for d in i])

        else:
            for i in self._data:
                if isinstance(i, dict):
                    for d in i:
                        i[d] = process_content(i[d])
                    json_data = i
                else:
                    json_data = [process_content(d) for d in i]

        with open(self.path, 'w', encoding=self.encoding) as f:
            dump(json_data, f)


def _set_style(_col_height, _row_styles, ws, recorder):
    if _col_height is not None:
        ws.row_dimensions[ws.max_row].height = recorder._col_height

    if _row_styles:
        groups = zip(ws[ws.max_row], _row_styles)
        for g in groups:
            g[1].to_cell(g[0])

    elif recorder._style:
        for c in ws[ws.max_row]:
            recorder._style.to_cell(c)


def _get_title(data: Union[list, dict],
               before: Union[list, dict, None] = None,
               after: Union[list, dict, None] = None) -> Union[list, None]:
    """获取表头列表
    :param data: 数据列表或字典
    :param before: 数据前的列
    :param after: 数据后的列
    :return: 表头列表
    """
    if isinstance(data, (tuple, list)):
        return None

    return_list = []
    for i in (before, data, after):
        if isinstance(i, dict):
            return_list.extend(list(i))
        elif i is None:
            pass
        elif isinstance(i, list):
            return_list.extend(['' for _ in range(len(i))])
        else:
            return_list.extend([''])

    return return_list
